#!/usr/bin/env python3

import argparse
import json
import re
import subprocess
from datetime import date
from io import BytesIO
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook

DEFAULT_SOURCE_URL = "https://csrc.nist.gov/files/pubs/sp/800/161/r1/upd1/final/docs/sp800-161r1-table26.xlsx"
DEFAULT_SOURCE_PAGE = "https://csrc.nist.gov/pubs/sp/800/161/r1/upd1/final"
DEFAULT_SOURCE_VERSION = "NIST SP 800-161 Rev.1 Update 1 (2024-11-01)"

SECTION_RE = re.compile(r"^Section\s+([0-9]+):\s*(.+)$", re.IGNORECASE)


def fetch_workbook(source_url: str):
    payload = None
    try:
        req = Request(source_url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=60) as response:
            payload = response.read()
    except Exception:
        result = subprocess.run(
            [
                "curl",
                "-fsSL",
                "--retry",
                "5",
                "--retry-all-errors",
                "--retry-delay",
                "2",
                "--connect-timeout",
                "20",
                "--max-time",
                "180",
                source_url,
            ],
            check=True,
            capture_output=True,
        )
        payload = result.stdout

    if not payload:
        raise RuntimeError("Downloaded NIST 800-161 workbook is empty.")

    return load_workbook(filename=BytesIO(payload), data_only=True)


def normalize_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def slugify(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def parse_questionnaire(ws):
    sections = {}
    section_order = []
    current_section = None
    first_section_one_seen = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        col_a = normalize_text(row[0] if len(row) > 0 else None)
        col_c = normalize_text(row[2] if len(row) > 2 else None)
        if not col_a:
            continue

        section_match = SECTION_RE.match(col_a)
        if section_match:
            section_num = int(section_match.group(1))
            section_name = section_match.group(2).strip()
            # The official sheet repeats "Section 1: Request Overview" as a page header.
            # Keep parsing in the current section for these repeated rows.
            if section_num == 1 and first_section_one_seen:
                continue

            first_section_one_seen = first_section_one_seen or section_num == 1
            current_section = {
                "num": section_num,
                "name": section_name,
            }
            section_key = (section_num, section_name)
            if section_key not in sections:
                sections[section_key] = []
                section_order.append(section_key)
            continue

        if current_section is None:
            continue

        lower = col_a.lower()
        if lower.startswith("supply chain risk management assessment scoping questionnaire"):
            continue
        if lower in {"provide response", "response provided by:", "response provided by"}:
            continue

        section_key = (current_section["num"], current_section["name"])
        sections[section_key].append(
            {
                "question": col_a,
                "owner": col_c,
            }
        )

    # Deduplicate exact question strings inside each section while preserving order.
    for key in section_order:
        deduped = []
        seen = set()
        for item in sections[key]:
            q = item["question"]
            if q in seen:
                continue
            seen.add(q)
            deduped.append(item)
        sections[key] = deduped

    return section_order, sections


def build_framework(section_order, sections, source_url):
    categories = []
    total_requirements = 0
    total_subcategories = 0
    today = date.today().isoformat()

    for section_num, section_name in section_order:
        items = sections[(section_num, section_name)]
        reqs = []
        for index, item in enumerate(items, start=1):
            code = f"T26.S{section_num}.{index}"
            req_id = f"t26-s{section_num}-{index}"
            owner = item.get("owner") or "N/A"
            reqs.append(
                {
                    "id": req_id,
                    "code": code,
                    "name": item["question"],
                    "description": f"{item['question']}\n\nResponse provided by: {owner}",
                    "level": "1",
                    "verification": "Review documented questionnaire response and supporting evidence.",
                }
            )

        subcategory_code = f"T26.S{section_num}"
        subcategory_id = f"table26-s{section_num}"
        subcategory = {
            "id": subcategory_id,
            "code": subcategory_code,
            "name": f"Section {section_num}: {section_name}",
            "description": "Questions extracted from NIST SP 800-161 Rev.1 Update 1 Table 26.",
            "requirements": reqs,
        }
        category = {
            "id": f"section-{section_num}",
            "code": f"S{section_num}",
            "name": f"Section {section_num}: {section_name}",
            "description": "C-SCRM assessment scoping questionnaire section.",
            "requirements": len(reqs),
            "subcategories": [subcategory],
        }
        categories.append(category)
        total_requirements += len(reqs)
        total_subcategories += 1

    return {
        "id": "nist-800-161",
        "name": "NIST SP 800-161",
        "fullName": "NIST SP 800-161 Rev.1 Update 1 Cybersecurity Supply Chain Risk Management",
        "version": "Rev.1 Update 1",
        "type": "standard",
        "domain": "supply-chain-security",
        "description": (
            "NIST SP 800-161 Rev.1 Update 1 C-SCRM assessment scoping questionnaire "
            "(Table 26) extracted from the official machine-readable workbook."
        ),
        "website": DEFAULT_SOURCE_PAGE,
        "organization": "NIST",
        "releaseDate": "2024-11-01",
        "lastUpdated": today,
        "language": "en",
        "source": {
            "type": "official-machine-readable",
            "url": source_url,
            "documentPage": DEFAULT_SOURCE_PAGE,
            "version": DEFAULT_SOURCE_VERSION,
        },
        "stats": {
            "totalRequirements": total_requirements,
            "totalCategories": len(categories),
            "totalSubcategories": total_subcategories,
            "level1": total_requirements,
            "level2": 0,
            "level3": 0,
        },
        "categories": categories,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Pull NIST SP 800-161 Rev.1 Update 1 Table 26 from official NIST workbook."
    )
    parser.add_argument("--source-url", default=DEFAULT_SOURCE_URL)
    parser.add_argument(
        "--output",
        default="public/data/frameworks/nist-800-161-en.json",
        help="Path to output English framework JSON.",
    )
    parser.add_argument(
        "--report",
        default="docs/framework-checkpoints/nist-800-161-pull-report.json",
        help="Path to write pull report summary JSON.",
    )
    args = parser.parse_args()

    wb = fetch_workbook(args.source_url)
    ws = wb["Questionnaire"]
    section_order, sections = parse_questionnaire(ws)
    framework = build_framework(section_order, sections, args.source_url)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(framework, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    report = {
        "frameworkId": framework["id"],
        "sourceUrl": args.source_url,
        "sourcePage": DEFAULT_SOURCE_PAGE,
        "categories": framework["stats"]["totalCategories"],
        "subcategories": framework["stats"]["totalSubcategories"],
        "requirements": framework["stats"]["totalRequirements"],
        "output": str(output_path),
    }
    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(
        f"Pulled nist-800-161: categories={report['categories']} "
        f"subcategories={report['subcategories']} requirements={report['requirements']}"
    )
    print(f"Output: {report['output']}")
    print(f"Report: {report_path}")


if __name__ == "__main__":
    main()
