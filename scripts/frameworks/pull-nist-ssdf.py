#!/usr/bin/env python3

import argparse
import json
import re
import subprocess
from datetime import date
from io import BytesIO
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook

DEFAULT_SOURCE_URL = (
    "https://csrc.nist.gov/files/pubs/sp/800/218/final/docs/"
    "nist.sp.800-218.ssdf-table.xlsx"
)

GROUP_RE = re.compile(r"^(.*?)\s*\(([A-Z]{2})\):\s*(.+)$")
PRACTICE_RE = re.compile(r"^(.*?)\s*\(([A-Z]{2}\.\d+)\):\s*(.+)$")
TASK_RE = re.compile(r"^([A-Z]{2}\.\d+\.\d+):\s*(.+)$")

GROUP_ORDER = ["PO", "PS", "PW", "RV"]


def fetch_workbook(source_url: str):
    content = None
    try:
        req = Request(source_url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=60) as response:
            content = response.read()
    except Exception:
        # Some TLS stacks intermittently fail against NIST/Cloudflare.
        # curl is more resilient in this environment, so use it as fallback.
        result = subprocess.run(
            ["curl", "-sSL", source_url],
            check=True,
            capture_output=True,
        )
        content = result.stdout

    if not content:
        raise RuntimeError("Downloaded workbook is empty.")

    return load_workbook(filename=BytesIO(content), data_only=True)


def normalize_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_groups(ws_groups):
    groups = {}
    for row in ws_groups.iter_rows(min_row=1, max_col=1, values_only=True):
        text = normalize_text(row[0])
        if not text:
            continue
        match = GROUP_RE.match(text)
        if not match:
            continue

        name = match.group(1).strip()
        code = match.group(2).strip()
        description = match.group(3).strip()
        groups[code] = {
            "id": code.lower(),
            "code": code,
            "name": name,
            "description": description,
            "requirements": 0,
            "subcategories": [],
        }
    return groups


def parse_tasks(ws_ssdf):
    practices = {}
    current_practice_code = None

    for row in ws_ssdf.iter_rows(min_row=2, max_col=4, values_only=True):
        practice_cell, task_cell, examples_cell, references_cell = row

        practice_text = normalize_text(practice_cell)
        if practice_text:
            practice_match = PRACTICE_RE.match(practice_text)
            if practice_match:
                practice_name = practice_match.group(1).strip()
                practice_code = practice_match.group(2).strip()
                practice_description = practice_match.group(3).strip()
                current_practice_code = practice_code
                practices.setdefault(
                    practice_code,
                    {
                        "id": practice_code.lower().replace(".", "-"),
                        "code": practice_code,
                        "name": practice_name,
                        "description": practice_description,
                        "requirements": [],
                    },
                )

        task_text = normalize_text(task_cell)
        if not task_text:
            continue

        task_match = TASK_RE.match(task_text)
        if not task_match:
            continue

        task_code = task_match.group(1).strip()
        task_name = task_match.group(2).strip()
        practice_code_from_task = ".".join(task_code.split(".")[:2])

        if current_practice_code is None:
            current_practice_code = practice_code_from_task

        target_practice = practices.setdefault(
            practice_code_from_task,
            {
                "id": practice_code_from_task.lower().replace(".", "-"),
                "code": practice_code_from_task,
                "name": practice_code_from_task,
                "description": "",
                "requirements": [],
            },
        )

        examples = normalize_text(examples_cell)
        references = normalize_text(references_cell)

        description_parts = [task_name]
        if examples:
            description_parts.append(f"Examples: {examples}")
        if references:
            description_parts.append(f"References: {references}")

        target_practice["requirements"].append(
            {
                "id": task_code.lower().replace(".", "-"),
                "code": task_code,
                "name": task_name,
                "description": "\n\n".join(description_parts),
                "level": "1",
                "verification": "Review documented implementation evidence for this SSDF task.",
            }
        )

    return practices


def build_framework(groups, practices, source_url):
    categories = []

    for group_code in GROUP_ORDER:
        group = groups.get(group_code)
        if not group:
            continue

        group_practices = [
            practice
            for practice_code, practice in practices.items()
            if practice_code.startswith(f"{group_code}.")
        ]
        group_practices.sort(
            key=lambda p: int(p["code"].split(".")[1]) if "." in p["code"] else 999
        )

        subcategories = []
        total_requirements = 0

        for practice in group_practices:
            requirement_count = len(practice["requirements"])
            total_requirements += requirement_count
            subcategories.append(
                {
                    "id": practice["id"],
                    "code": practice["code"],
                    "name": practice["name"],
                    "description": practice["description"],
                    "requirements": practice["requirements"],
                }
            )

        category = dict(group)
        category["requirements"] = total_requirements
        category["subcategories"] = subcategories
        categories.append(category)

    total_subcategories = sum(len(c["subcategories"]) for c in categories)
    total_requirements = sum(
        len(sub["requirements"])
        for category in categories
        for sub in category["subcategories"]
    )

    today = date.today().isoformat()
    framework = {
        "id": "nist-ssdf",
        "name": "NIST SSDF",
        "fullName": "NIST Secure Software Development Framework",
        "version": "1.1",
        "type": "standard",
        "domain": "application-security",
        "description": (
            "NIST SSDF v1.1 extracted from the official machine-readable "
            "SSDF table workbook."
        ),
        "website": "https://csrc.nist.gov/projects/ssdf",
        "organization": "NIST",
        "releaseDate": "2022-02-03",
        "lastUpdated": today,
        "language": "en",
        "source": {
            "type": "official-machine-readable",
            "url": source_url,
        },
        "stats": {
            "totalRequirements": total_requirements,
            "totalCategories": len(categories),
            "totalSubcategories": total_subcategories,
            "level1": total_requirements,
            "level2": 0,
            "level3": 0,
        },
        "categories": categories,
    }
    return framework


def main():
    parser = argparse.ArgumentParser(
        description="Pull NIST SSDF from official NIST SSDF table workbook."
    )
    parser.add_argument("--source-url", default=DEFAULT_SOURCE_URL)
    parser.add_argument(
        "--output",
        default="public/data/frameworks/nist-ssdf-en.json",
        help="Path to output English framework JSON.",
    )
    parser.add_argument(
        "--report",
        default="docs/framework-checkpoints/nist-ssdf-pull-report.json",
        help="Path to write pull report summary JSON.",
    )
    args = parser.parse_args()

    wb = fetch_workbook(args.source_url)
    ws_groups = wb["Groups"]
    ws_ssdf = wb["SSDF"]

    groups = parse_groups(ws_groups)
    practices = parse_tasks(ws_ssdf)
    framework = build_framework(groups, practices, args.source_url)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(framework, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    report = {
        "frameworkId": framework["id"],
        "sourceUrl": args.source_url,
        "categories": framework["stats"]["totalCategories"],
        "subcategories": framework["stats"]["totalSubcategories"],
        "requirements": framework["stats"]["totalRequirements"],
        "output": str(output_path),
    }

    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(
        f"Pulled nist-ssdf: categories={report['categories']} "
        f"subcategories={report['subcategories']} requirements={report['requirements']}"
    )
    print(f"Output: {report['output']}")
    print(f"Report: {report_path}")


if __name__ == "__main__":
    main()
